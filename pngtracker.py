import requests
from datetime import datetime
import os
import json
import smtplib
from email.mime.text import MIMEText
from openpyxl import Workbook, load_workbook
import pytz

# -------------------------------
# 🇮🇳 IST TIMEZONE (FIXED)
# -------------------------------
IST = pytz.timezone("Asia/Kolkata")

# -------------------------------
# 📂 FILE PATH
# -------------------------------
EXCEL_FILE = "metal_rates.xlsx"
HISTORY_FILE = "last_prices.json"

# -------------------------------
# 📊 HEADERS
# -------------------------------
HEADERS_EXCEL = [
    "Date", "Time",
    "Gold 24K",
    "Gold 24K 995",
    "Gold 24K 995GW",
    "Gold 22K",
    "Gold 18K",
    "Gold 14K",
    "Gold 9K",
    "Silver",
    "Silver Bar",
    "Platinum",
    "Source"
]

METALS = [
    "Gold 24K",
    "Gold 24K 995",
    "Gold 24K 995GW",
    "Gold 22K",
    "Gold 18K",
    "Gold 14K",
    "Gold 9K",
    "Silver",
    "Silver Bar",
    "Platinum"
]

# -------------------------------
# 🔗 API
# -------------------------------
API_URL = "https://goldpriceeditor.droidinfinity.com/api/external/metal-prices/1085"

HEADERS = {
    "accept": "application/json",
    "origin": "https://pngadgilandsons.com",
    "referer": "https://pngadgilandsons.com/",
    "user-agent": "Mozilla/5.0"
}

# -------------------------------
# 📧 EMAIL FUNCTION
# -------------------------------
def send_email(subject, body):
    sender = os.getenv("EMAIL_USER")
    password = os.getenv("EMAIL_PASS")

    if not sender or not password:
        print("No email credentials")
        return

    password = password.replace(" ", "")

    msg = MIMEText(body)
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = sender

    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(sender, password)
        server.send_message(msg)
        server.quit()
        print("📧 Email Sent")
    except Exception as e:
        print("Email Error:", e)

# -------------------------------
# 📂 HISTORY
# -------------------------------
def load_last_prices():
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, "r") as f:
            return json.load(f)
    return {}

def save_last_prices(data):
    with open(HISTORY_FILE, "w") as f:
        json.dump(data, f)

# -------------------------------
# ✅ GET DATA (FIXED)
# -------------------------------
def get_rates():
    try:
        res = requests.get(API_URL, headers=HEADERS, timeout=10)

        if res.status_code != 200:
            print("❌ API failed:", res.status_code)
            return None

        data_json = res.json()

        if "rates" not in data_json:
            print("❌ Invalid API response")
            return None

        d = data_json["rates"]

        data = {
            "Gold 24K": d.get("goldPrice24K"),
            "Gold 24K 995": d.get("goldPrice24K995"),
            "Gold 24K 995GW": d.get("goldPrice24K995GW"),
            "Gold 22K": d.get("goldPrice22K"),
            "Gold 18K": d.get("goldPrice18K"),
            "Gold 14K": d.get("goldPrice14K"),
            "Gold 9K": d.get("goldPrice9K"),
            "Silver": d.get("silverPrice"),
            "Silver Bar": d.get("silverBarPrice"),
            "Platinum": d.get("platinumPrice"),
            "Source": "API"
        }

        print("📊 Data fetched:", data)  # Debug

        return data

    except Exception as e:
        print("❌ API Error:", e)
        return None

# -------------------------------
# 📊 SAVE TO EXCEL (IST FIXED)
# -------------------------------
def save_excel(data):
    # ✅ Correct IST conversion
    now = datetime.now(pytz.utc).astimezone(IST)

    date = now.strftime("%Y-%m-%d")
    time_now = now.strftime("%H:%M:%S")

    print("📍 Saving to:", EXCEL_FILE)

    # Create file if not exists
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS_EXCEL)
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Fix header mismatch
    current_headers = [cell.value for cell in ws[1]]
    if current_headers != HEADERS_EXCEL:
        print("⚠ Fixing header mismatch...")
        ws.delete_rows(1, ws.max_row)
        ws.append(HEADERS_EXCEL)
        wb.save(EXCEL_FILE)

    # Add row
    row = [
        date,
        time_now,
        data.get("Gold 24K"),
        data.get("Gold 24K 995"),
        data.get("Gold 24K 995GW"),
        data.get("Gold 22K"),
        data.get("Gold 18K"),
        data.get("Gold 14K"),
        data.get("Gold 9K"),
        data.get("Silver"),
        data.get("Silver Bar"),
        data.get("Platinum"),
        data.get("Source")
    ]

    ws.append(row)
    wb.save(EXCEL_FILE)

    print("✅ Data saved correctly")

# -------------------------------
# 🎯 MAIN
# -------------------------------
def main():
    print("\n⏳ Running at:", datetime.now(pytz.utc).astimezone(IST).strftime("%H:%M:%S"))

    data = get_rates()
    if not data:
        return

    save_excel(data)

    last = load_last_prices()

    if last:
        changes = []
        for m in METALS:
            if str(last.get(m)) != str(data.get(m)):
                changes.append(f"{m}: {last.get(m)} → {data.get(m)}")

        if changes:
            body = "Price Changes:\n\n" + "\n".join(changes)
            send_email("Gold Price Alert 🚨", body)

    save_last_prices(data)

# -------------------------------
# ▶ RUN
# -------------------------------
if __name__ == "__main__":
    main()
